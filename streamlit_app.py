#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI 驱动的定价与补贴优化系统 — Streamlit 看板 v2
"""

import streamlit as st
import json
from collections import defaultdict

st.set_page_config(page_title="AI 补贴策略仿真平台", page_icon="🎯", layout="wide", initial_sidebar_state="collapsed")

# ═══ 全局样式 ═══
st.markdown("""
<style>
    .main > div { padding-top: 0.5rem; }
    [data-testid="stSidebar"] { display: none; }
    div[data-testid="stTabs"] button { font-size: 14px !important; font-weight: 500 !important; padding: 8px 20px !important; }
    div[data-testid="stTabs"] { margin-bottom: 0; }
    .block-container { padding-top: 0 !important; padding-bottom: 1rem; }
    div[data-testid="stVerticalBlock"] > div { gap: 0.3rem !important; }
    div[data-testid="stExpander"] { margin: 0 !important; }
    .stMarkdown { margin: 0 !important; padding: 0 !important; }
    .card { margin-top: 0.3rem !important; margin-bottom: 0.3rem !important; }
    h1 { font-size: 1.6rem !important; color: #1a1a2e; }
    h2 { font-size: 1.2rem !important; color: #333; }
    h3 { font-size: 1.05rem !important; }
    .kpi-box { background: white; border-radius: 10px; padding: 18px 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); text-align: center; }
    .kpi-box .label { font-size: 12px; color: #888; margin-bottom: 4px; }
    .kpi-box .value { font-size: 26px; font-weight: 700; }
    .kpi-box .sub { font-size: 11px; color: #aaa; margin-top: 2px; }
    .card { background: white; border-radius: 10px; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); margin-bottom: 16px; }
    .alert-red { border-left: 4px solid #e53935; background: #fff5f5; border-radius: 8px; padding: 14px 18px; margin-bottom: 12px; }
    .alert-orange { border-left: 4px solid #f57c00; background: #fff8e1; border-radius: 8px; padding: 14px 18px; margin-bottom: 12px; }
    .alert-green { border-left: 4px solid #2e7d32; background: #f1f8e9; border-radius: 8px; padding: 14px 18px; margin-bottom: 12px; }
    .strategy-table { width: 100%; border-collapse: collapse; font-size: 13px; }
    .strategy-table th { background: #1a1a2e; color: white; padding: 10px 12px; text-align: center; }
    .strategy-table td { padding: 10px 12px; text-align: center; border-bottom: 1px solid #eee; }
    .strategy-table tr:hover { background: #f5f5f5; }
    .tag-green { background: #e8f5e9; color: #2e7d32; padding: 3px 10px; border-radius: 12px; font-size: 12px; font-weight: 600; }
    .tag-red { background: #ffebee; color: #c62828; padding: 3px 10px; border-radius: 12px; font-size: 12px; font-weight: 600; }
    .tag-yellow { background: #fff8e1; color: #e65100; padding: 3px 10px; border-radius: 12px; font-size: 12px; font-weight: 600; }
    .funnel-step { text-align: center; padding: 14px 0; }
    .funnel-bar { margin: 0 auto; height: 48px; border-radius: 6px; display: flex; align-items: center; justify-content: space-between; padding: 0 16px; color: white; font-weight: 600; }
    .funnel-arrow { text-align: center; font-size: 12px; padding: 4px 0; }
</style>
""", unsafe_allow_html=True)

# ═══ 数据加载 ═══
@st.cache_data
def load_phase2():
    with open("phase2_calibrated_final.json", "r", encoding="utf-8") as f:
        return json.load(f)

@st.cache_data
def load_user_features():
    import openpyxl
    wb = openpyxl.load_workbook("user_simulation_features.xlsx")
    ws = wb["Sheet1"]
    headers = [cell.value for cell in ws[1]]
    users = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        d = dict(zip(headers, row))
        users[d["user_id"]] = d
    return users

SCENARIOS = {
    "weekday_lunch": {"label": "工作日午餐", "icon": "🍱", "coupon": "满30减8", "amount": 8, "threshold": 30},
    "afternoon_tea": {"label": "下午茶", "icon": "🧋", "coupon": "满25减8", "amount": 8, "threshold": 25},
    "weekend_rain_dinner": {"label": "周末暴雨晚餐", "icon": "🌧️", "coupon": "满50减15", "amount": 15, "threshold": 50},
    "weekday_breakfast": {"label": "工作日早餐", "icon": "🥐", "coupon": "满15减5", "amount": 5, "threshold": 15},
    "weekend_no_coupon": {"label": "周末无券(对照)", "icon": "🚫", "coupon": "无", "amount": 0, "threshold": 0},
    "late_night": {"label": "夜宵", "icon": "🌙", "coupon": "满20减6", "amount": 6, "threshold": 20},
    "holiday_gathering": {"label": "节假日聚餐", "icon": "🎉", "coupon": "满80减25", "amount": 25, "threshold": 80},
    "rain_stuck": {"label": "暴雨困办公室", "icon": "⛈️", "coupon": "满30减12", "amount": 12, "threshold": 30},
}
CONTROL = "weekend_no_coupon"

phase2 = load_phase2()
user_features = load_user_features()
user_scenarios = defaultdict(dict)
for t in phase2:
    user_scenarios[t["user_id"]][t["scenario_key"]] = t

# ═══ 顶部导航 ═══
st.markdown("""
<style>
    header[data-testid="stHeader"] { display: none; }
    .block-container { padding-top: 0 !important; }
    .top-bar {
        background: linear-gradient(135deg, #0d1b3e, #1a3a6e);
        padding: 22px 32px 18px 32px;
        margin: -1rem -4rem 0.5rem -4rem;
    }
    .top-bar h1 { font-size: 20px; font-weight: 700; color: white; margin: 0; letter-spacing: 0.5px; }
    .top-bar p { font-size: 12px; color: rgba(255,255,255,0.55); margin: 4px 0 0 0; }
</style>
<div class="top-bar">
    <h1>基于 AI Agent 的补贴策略仿真与优化系统</h1>
    <p>仿真校准 SiFi 0.950 (A级) · 200 用户 × 8 场景 · 下单率偏差 3.6% · 数据周期 2026-03 ~ 2026-04</p>
</div>
""", unsafe_allow_html=True)

# 顶部 Tab 导航
tab_overview, tab_sim, tab_user, tab_realtime, tab_risk = st.tabs(["📊 系统总览", "🔬 策略仿真", "👤 用户模拟", "🧪 实时仿真", "⚠️ 风险与建议"])

# ═══════════════════════════════════════════════════════════════
# 页面 1: 系统总览
# ═══════════════════════════════════════════════════════════════
with tab_overview:

    # 方法论流程
    st.markdown("#### 🔬 技术架构与方法论")
    st.markdown("""
    <div style="display:flex;align-items:center;justify-content:center;gap:8px;flex-wrap:wrap;padding:12px 0">
        <div style="background:#e3f2fd;padding:10px 16px;border-radius:8px;text-align:center;min-width:130px">
            <div style="font-size:18px">📊</div>
            <div style="font-size:12px;font-weight:600;color:#1565c0">Step 1</div>
            <div style="font-size:11px;color:#555">特征工程</div>
            <div style="font-size:10px;color:#999">36维用户画像</div>
        </div>
        <div style="color:#999;font-size:18px">→</div>
        <div style="background:#e8f5e9;padding:10px 16px;border-radius:8px;text-align:center;min-width:130px">
            <div style="font-size:18px">🧠</div>
            <div style="font-size:12px;font-weight:600;color:#2e7d32">Step 2</div>
            <div style="font-size:11px;color:#555">Persona 注入</div>
            <div style="font-size:10px;color:#999">三模块 Prompt</div>
        </div>
        <div style="color:#999;font-size:18px">→</div>
        <div style="background:#fff3e0;padding:10px 16px;border-radius:8px;text-align:center;min-width:130px">
            <div style="font-size:18px">🤖</div>
            <div style="font-size:12px;font-weight:600;color:#e65100">Step 3</div>
            <div style="font-size:11px;color:#555">Agent 仿真</div>
            <div style="font-size:10px;color:#999">200人×8场景</div>
        </div>
        <div style="color:#999;font-size:18px">→</div>
        <div style="background:#fce4ec;padding:10px 16px;border-radius:8px;text-align:center;min-width:130px">
            <div style="font-size:18px">🔧</div>
            <div style="font-size:12px;font-weight:600;color:#c62828">Step 4</div>
            <div style="font-size:11px;color:#555">Prompt 校准</div>
            <div style="font-size:10px;color:#999">7条约束规则</div>
        </div>
        <div style="color:#999;font-size:18px">→</div>
        <div style="background:#e8eaf6;padding:10px 16px;border-radius:8px;text-align:center;min-width:130px">
            <div style="font-size:18px">✅</div>
            <div style="font-size:12px;font-weight:600;color:#283593">Step 5</div>
            <div style="font-size:11px;color:#555">SiFi 评估</div>
            <div style="font-size:10px;color:#999">四维保真度</div>
        </div>
        <div style="color:#999;font-size:18px">→</div>
        <div style="background:#e0f2f1;padding:10px 16px;border-radius:8px;text-align:center;min-width:130px">
            <div style="font-size:18px">🎯</div>
            <div style="font-size:12px;font-weight:600;color:#00695c">Step 6</div>
            <div style="font-size:11px;color:#555">策略优化</div>
            <div style="font-size:10px;color:#999">ECUP + 拉格朗日</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # 各步骤详情（可展开）
    st.markdown("#### 📋 各步骤详情")

    with st.expander("**Step 1 — 特征工程：** 解决 200 用户仅 8 条行为轨迹的冷启动问题", expanded=False):
        st.markdown("""
        - 三级画像感知合成引擎：Tier-1 订单反推行为漏斗 → Tier-2 画像属性高斯采样 → Tier-3 全局中位数兜底
        - 马尔可夫链扩充稀疏用户的状态转移序列
        - KL 散度校验合成质量（阈值 < 0.5）
        - 5 个并行模块提取 36 维特征矩阵（微观决策心智、动态价格弹性、时间衰减 RFM、社会属性标签等）
        """)

    with st.expander("**Step 2 — Persona 注入：** 将数值特征无损翻译为 LLM 可执行的行为规则", expanded=False):
        st.markdown("""
        - **Module A 身份内核：** 年龄/城市/会员/生命周期等标签 → 角色锚定
        - **Module B 决策规则：** `avg_basket_size=35` → "低于25元划算，25-46元舒适区，超过53元犹豫"
        - **Module C CoT 思维链：** 按弹性类型生成三种不同的决策路径
            - 薅羊毛党：券值评估 → 凑单计算 → 替代比较 → 决策
            - 品质用户：需求评估 → 品牌熟悉度 → 价格确认 → 决策
            - 刚需用户：需求判断 → 习惯选择 → 价格确认 → 决策
        """)

    with st.expander("**Step 3 — Agent 仿真：** 200 用户 × 8 场景全量仿真", expanded=False):
        st.markdown("""
        - 8 个控制变量实验场景（覆盖暴雨/下午茶/节假日/夜宵等）
        - 调用多个 LLM API（豆包/智谱/ModelScope）
        - 每个 Agent 输出：action（下单/加购/浏览/忽略）、confidence、CoT thinking
        """)

    with st.expander("**Step 4 — 仿真校准：** Prompt 校准 vs LoRA 微调双路径对比", expanded=True):
        st.markdown("**核心问题：** LLM 天然倾向于给出积极配合的回答，导致下单率虚高（84.7% vs 真实 47.8%）")
        st.markdown("")

        # 三阶段对比表
        st.markdown("""
        <table class="strategy-table">
            <tr>
                <th>指标</th>
                <th>Phase 1<br><span style="font-weight:400;font-size:11px">无校准</span></th>
                <th>Phase 2<br><span style="font-weight:400;font-size:11px">Prompt 校准</span></th>
                <th>Phase 3<br><span style="font-weight:400;font-size:11px">LoRA 微调</span></th>
                <th>真实数据</th>
            </tr>
            <tr>
                <td><b>整体下单率</b></td>
                <td><span class="tag-red">84.7%</span></td>
                <td><span class="tag-yellow">51.4%</span></td>
                <td><span class="tag-green">49.2%</span></td>
                <td>47.8%</td>
            </tr>
            <tr>
                <td><b>券使用率</b></td>
                <td><span class="tag-red">83.0%</span></td>
                <td><span class="tag-yellow">74.1%</span></td>
                <td><span class="tag-green">50.3%</span></td>
                <td>47.8%</td>
            </tr>
            <tr>
                <td><b>置信度均值</b></td>
                <td><span class="tag-red">0.868</span></td>
                <td><span class="tag-green">0.664</span></td>
                <td><span class="tag-green">0.612</span></td>
                <td>[0.45-0.75]</td>
            </tr>
            <tr>
                <td><b>置信度标准差</b></td>
                <td><span class="tag-red">0.081</span></td>
                <td><span class="tag-yellow">0.159</span></td>
                <td><span class="tag-green">0.218</span></td>
                <td>> 0.20</td>
            </tr>
            <tr>
                <td><b>弹性类型排序</b></td>
                <td><span class="tag-red">❌ 逆转</span></td>
                <td><span class="tag-green">✅ 正确</span></td>
                <td><span class="tag-green">✅ 正确</span></td>
                <td>刚需>品质>薅羊毛</td>
            </tr>
            <tr>
                <td><b>刚需下单率</b></td>
                <td>86.1%</td>
                <td>53.6%</td>
                <td>51.8%</td>
                <td>~52%</td>
            </tr>
            <tr>
                <td><b>品质下单率</b></td>
                <td>86.5%</td>
                <td>48.6%</td>
                <td>42.1%</td>
                <td>~40%</td>
            </tr>
            <tr>
                <td><b>薅羊毛党下单率</b></td>
                <td>56.9%</td>
                <td>26.3%</td>
                <td>22.7%</td>
                <td>~25%</td>
            </tr>
            <tr>
                <td><b>SiFi 总分</b></td>
                <td><span class="tag-red">0.340 (D级)</span></td>
                <td><span class="tag-green">0.950 (A级)</span></td>
                <td><span class="tag-green">0.978 (A级)</span></td>
                <td>≥ 0.85 为A级</td>
            </tr>
        </table>
        """, unsafe_allow_html=True)

        st.markdown("")

        # 两种方法的对比分析
        method1, method2 = st.columns(2)
        with method1:
            st.markdown("""
            <div class="alert-green">
                <strong>✅ Prompt 校准（7 条约束规则）</strong><br>
                <span style="font-size:12px">
                <b>优势：</b>零成本、即时生效、不需要 GPU<br>
                <b>效果：</b>下单率偏差 3.6%，SiFi 0.950<br>
                <b>局限：</b>券使用率仍偏高（74.1% vs 47.8%），置信度标准差偏小
                </span>
            </div>
            """, unsafe_allow_html=True)

        with method2:
            st.markdown("""
            <div class="alert-green">
                <strong>✅ LoRA 微调（RFT 筛选 + rank=16）</strong><br>
                <span style="font-size:12px">
                <b>优势：</b>券使用率精确对齐（50.3%），置信度分布更合理<br>
                <b>效果：</b>下单率偏差 1.4%，SiFi 0.978<br>
                <b>成本：</b>需要 GPU（A100 约 2-3 小时），6,400 条 RFT 筛选数据
                </span>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("""
        <div style="background:#f5f5f5;border-radius:8px;padding:14px 18px;margin-top:8px;font-size:13px">
            <b>结论：</b>Prompt 校准已达到 A 级（0.950），满足策略决策需求。LoRA 微调进一步将券使用率从 74.1% 校准至 50.3%，
            置信度标准差从 0.159 提升至 0.218，SiFi 从 0.950 提升至 0.978。两种方法互补：Prompt 校准解决方向性偏差，
            LoRA 微调解决精度问题。最终系统采用 LoRA 微调后的模型作为生产版本。
        </div>
        """, unsafe_allow_html=True)

    with st.expander("**Step 5 — SiFi 保真度评估：** 四维统计检验量化仿真可信度", expanded=False):
        sifi_data = [
            ("行为分布对齐", 0.30, 0.10, 1.00, "卡方检验 + 绝对偏差"),
            ("弹性类型差异度", 0.25, 0.20, 1.00, "排序检验 + 差异幅度"),
            ("场景响应合理性", 0.25, 0.80, 0.80, "5项因果方向检验"),
            ("置信度校准质量", 0.20, 0.30, 1.00, "均值/标准差/分组对比"),
        ]
        for name, weight, p1, p2, method in sifi_data:
            sc1, sc2, sc3, sc4 = st.columns([3, 1, 1, 1])
            sc1.caption(f"**{name}** (权重{weight}) — {method}")
            sc2.caption(f"Phase1: {p1}")
            sc3.caption(f"Phase2: {p2}")
            sc4.caption(f"{'✅' if p2 >= 0.8 else '⚠️'}")
        st.metric("SiFi 总分", "0.340 → 0.950", "+0.610 (D级→A级)")

    with st.expander("**Step 6 — 策略优化：** Agent 反事实 + ECUP 因果增益 + 拉格朗日寻优", expanded=False):
        st.markdown("""
        - **反事实对比：** 同一 Agent 在有券/无券场景分别决策，差值即因果增量
        - **ECUP 模型：** uplift_conv = P(下单|有券) - P(下单|无券)
        - **预算优化：** 拉格朗日对偶搜索，在预算约束下最大化增量 GMV + 转化
        - **劣质补贴识别：** 按 用户类型×场景 精细化识别 net_value < 0 的组合
        """)


    # KPI 概览
    st.markdown("#### 📈 核心产出指标")
    total = len(phase2)
    orders = sum(1 for t in phase2 if t["action"] == "order")
    order_rate = orders / total

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.markdown(f'<div class="kpi-box"><div class="label">增量 GMV</div><div class="value" style="color:#2e7d32">¥4,197</div><div class="sub">budget=300</div></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="kpi-box"><div class="label">ROI</div><div class="value" style="color:#1565c0">14.0x</div><div class="sub">每1元补贴回报</div></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="kpi-box"><div class="label">仿真转化率</div><div class="value">{order_rate:.1%}</div><div class="sub">真实 47.8%</div></div>', unsafe_allow_html=True)
    c4.markdown(f'<div class="kpi-box"><div class="label">正向占比</div><div class="value">69.1%</div><div class="sub">uplift > 0</div></div>', unsafe_allow_html=True)
    c5.markdown(f'<div class="kpi-box"><div class="label">SiFi 评级</div><div class="value" style="color:#2e7d32">A 级</div><div class="sub">0.950 / 1.000</div></div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# 页面 2: 策略仿真
# ═══════════════════════════════════════════════════════════════
with tab_sim:
    st.markdown("## 🔬 策略仿真 — 场景级 A/B 对比")

    scene_keys = [k for k in SCENARIOS if k != CONTROL]
    tabs = st.tabs([f"{SCENARIOS[k]['icon']} {SCENARIOS[k]['label']}" for k in scene_keys])

    for idx, sk in enumerate(scene_keys):
        with tabs[idx]:
            sc = SCENARIOS[sk]
            scene_data = [t for t in phase2 if t["scenario_key"] == sk]
            ctrl_data = [t for t in phase2 if t["scenario_key"] == CONTROL]

            s_orders = sum(1 for t in scene_data if t["action"] == "order")
            c_orders = sum(1 for t in ctrl_data if t["action"] == "order")
            s_rate = s_orders / max(len(scene_data), 1)
            c_rate = c_orders / max(len(ctrl_data), 1)
            uplift = (s_rate - c_rate) * 100

            # 补贴前后对比
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("无券转化率", f"{c_rate:.1%}")
            m2.metric("有券转化率", f"{s_rate:.1%}", f"+{uplift:.1f}pp")
            m3.metric("增量订单", f"+{s_orders - c_orders}", f"共{len(scene_data)}人")
            m4.metric("券面额", f"¥{sc['amount']}", f"门槛 ¥{sc['threshold']}")

            # 弹性拆分 + 行为分布
            el, er = st.columns(2)
            with el:
                st.markdown("**按弹性类型拆分**")
                for etype in ["缺乏弹性(刚需)", "低弹性(撇脂定价目标)", "高弹性(薅羊毛党)"]:
                    sub_s = [t for t in scene_data if t["elasticity_type"] == etype]
                    sub_c = [t for t in ctrl_data if t["elasticity_type"] == etype]
                    if not sub_s: continue
                    sr = sum(1 for t in sub_s if t["action"] == "order") / len(sub_s)
                    cr = sum(1 for t in sub_c if t["action"] == "order") / max(len(sub_c), 1)
                    d = (sr - cr) * 100
                    short = etype.split("(")[0]
                    st.progress(sr, text=f"{short}: **{sr:.1%}** (uplift +{d:.1f}pp)")

            with er:
                st.markdown("**行为分布**")
                acts = defaultdict(int)
                for t in scene_data: acts[t["action"]] += 1
                total_s = len(scene_data)
                for a, emoji in [("order","✅"),("cart","🛒"),("browse","👀"),("ignore","💤")]:
                    cnt = acts.get(a, 0)
                    pct = cnt / max(total_s, 1)
                    st.progress(pct, text=f"{emoji} {a}: {cnt} ({pct:.1%})")

# ═══════════════════════════════════════════════════════════════
# 页面 3: 用户模拟
# ═══════════════════════════════════════════════════════════════
with tab_user:
    st.markdown("## 👤 用户模拟 — 单用户决策全景可视化")

    type_users = defaultdict(list)
    for uid in sorted(user_scenarios.keys()):
        first = list(user_scenarios[uid].values())[0]
        type_users[first["elasticity_type"]].append(uid)

    col_sel, col_detail = st.columns([1, 3])

    with col_sel:
        st.markdown("#### 选择用户")
        sel_type = st.selectbox("弹性类型筛选", list(type_users.keys()), key="user_type_sel")
        sel_user = st.selectbox("用户 ID", type_users[sel_type][:30], key="user_id_sel")

        # 字段说明卡片
        st.markdown("#### 📖 字段说明")
        st.markdown("""
        <div style="font-size:12px;line-height:1.8">
        <b>客单价</b>：用户历史平均每单消费金额<br>
        <b>弹性类型</b>：对价格变化的敏感程度<br>
        &nbsp;&nbsp;· 刚需：饿了就买，券影响小<br>
        &nbsp;&nbsp;· 品质导向：看品质不看价格<br>
        &nbsp;&nbsp;· 薅羊毛党：有券才买<br>
        <b>冲动下单率</b>：看到即买的概率<br>
        <b>深度比价率</b>：多店比较的概率<br>
        <b>品牌集中度(HHI)</b>：越高越忠诚于固定店铺<br>
        <b>活跃健康度</b>：近期活跃程度(0-1)<br>
        <b>置信度</b>：Agent 对决策的确信程度
        </div>
        """, unsafe_allow_html=True)

    with col_detail:
        if sel_user:
            feat = user_features.get(sel_user, {})
            scenes = user_scenarios[sel_user]

            # ── 用户画像卡片 ──
            st.markdown(f"#### 🧑 {sel_user} 用户画像")

            # 核心指标
            k1, k2, k3, k4, k5 = st.columns(5)
            k1.markdown(f'<div class="kpi-box"><div class="label">客单价</div><div class="value">¥{float(feat.get("avg_basket_size",0)):.0f}</div></div>', unsafe_allow_html=True)
            k2.markdown(f'<div class="kpi-box"><div class="label">弹性类型</div><div class="value" style="font-size:16px">{str(feat.get("elasticity_type",""))}</div></div>', unsafe_allow_html=True)
            k3.markdown(f'<div class="kpi-box"><div class="label">活跃健康度</div><div class="value">{float(feat.get("lifecycle_health_score",0)):.2f}</div></div>', unsafe_allow_html=True)
            k4.markdown(f'<div class="kpi-box"><div class="label">品牌集中度</div><div class="value">{float(feat.get("brand_hhi",0)):.2f}</div></div>', unsafe_allow_html=True)
            basket_val = float(feat.get("avg_basket_size", 40))
            k5.markdown(f'<div class="kpi-box"><div class="label">价格舒适区</div><div class="value" style="font-size:14px">¥{int(basket_val*0.7)}-{int(basket_val*1.3)}</div></div>', unsafe_allow_html=True)

            st.markdown("")

            # 行为特征 + 标签
            bf1, bf2 = st.columns(2)
            with bf1:
                st.markdown("**📊 行为特征**")
                imp = float(feat.get("impulse_buy_rate", 0))
                cmp = float(feat.get("deep_compare_rate", 0))
                crt = float(feat.get("avg_add_cart_per_session", 0))
                ent = float(feat.get("category_entropy", 0))
                ln = int(feat.get("late_night_actions", 0))
                bw = float(feat.get("bad_weather_order_ratio", 0))

                st.progress(min(imp, 1.0), text=f"冲动下单率: **{imp*100:.0f}%** {'(高冲动)' if imp > 0.15 else '(低冲动)' if imp < 0.08 else ''}")
                st.progress(min(cmp, 1.0), text=f"深度比价率: **{cmp*100:.0f}%** {'(爱比价)' if cmp > 0.2 else '(很少比价)' if cmp < 0.1 else ''}")
                st.progress(min(crt/2, 1.0), text=f"加购频次: **{crt:.1f}次/会话** {'(重度加购)' if crt > 1.0 else '(很少加购)' if crt < 0.3 else ''}")
                st.progress(min(ent/5, 1.0), text=f"品类探索广度: **{ent:.1f}** {'(广泛)' if ent > 3.5 else '(集中)' if ent < 2.5 else '(适中)'}")

            with bf2:
                st.markdown("**🏷️ 用户标签**")
                tag_data = [
                    ("心理类型", feat.get("psychology_tag", "未知"), "用户的消费决策风格"),
                    ("品牌态度", feat.get("loyalty_tag", "未知"), "对品牌的忠诚程度"),
                    ("天气响应", feat.get("weather_tag", "未知"), "恶劣天气对下单的影响"),
                    ("家庭状态", feat.get("household_tag", "未知"), "单人/多人用餐习惯"),
                ]
                for label, value, desc in tag_data:
                    st.markdown(f"""
                    <div style="background:#f8f9fa;border-radius:6px;padding:8px 12px;margin-bottom:6px">
                        <span style="font-size:12px;color:#888">{label}</span>
                        <span style="font-size:11px;color:#bbb;margin-left:4px">({desc})</span><br>
                        <span style="font-size:14px;font-weight:600">{value}</span>
                    </div>
                    """, unsafe_allow_html=True)

                st.markdown(f"""
                <div style="background:#f8f9fa;border-radius:6px;padding:8px 12px;margin-bottom:6px">
                    <span style="font-size:12px;color:#888">场景敏感度</span><br>
                    <span style="font-size:13px">🌙 夜宵行为 <b>{ln}次</b> {'(有夜宵习惯)' if ln > 2 else '(无夜宵习惯)' if ln == 0 else ''}</span><br>
                    <span style="font-size:13px">🌧️ 恶劣天气 <b>{bw*100:.0f}%</b> {'(重度依赖)' if bw > 0.3 else '(不敏感)' if bw < 0.1 else ''}</span>
                </div>
                """, unsafe_allow_html=True)


            # ── 场景决策 ──
            st.markdown("#### 🎯 场景决策详情")
            sel_scene = st.selectbox("选择场景查看决策过程", list(scenes.keys()), key="scene_sel",
                                     format_func=lambda x: f"{SCENARIOS.get(x,{}).get('icon','')} {SCENARIOS.get(x,{}).get('label',x)} ({SCENARIOS.get(x,{}).get('coupon','')})")
            traj = scenes[sel_scene]

            # 决策结果
            action_map = {"order": ("✅ 下单", "#2e7d32", "#e8f5e9"), "cart": ("🛒 加购", "#e65100", "#fff3e0"),
                          "browse": ("👀 浏览", "#1565c0", "#e3f2fd"), "ignore": ("💤 忽略", "#999", "#f5f5f5")}
            a_text, a_color, a_bg = action_map.get(traj["action"], ("未知", "#999", "#f5f5f5"))

            st.markdown(f"""
            <div style="display:flex;gap:16px;margin:12px 0">
                <div style="background:{a_bg};border:2px solid {a_color};border-radius:10px;padding:16px 24px;text-align:center;flex:1">
                    <div style="font-size:12px;color:#888">决策结果</div>
                    <div style="font-size:22px;font-weight:700;color:{a_color}">{a_text}</div>
                </div>
                <div style="background:#f8f9fa;border-radius:10px;padding:16px 24px;text-align:center;flex:1">
                    <div style="font-size:12px;color:#888">置信度</div>
                    <div style="font-size:22px;font-weight:700">{traj['confidence']:.2f}</div>
                </div>
                <div style="background:#f8f9fa;border-radius:10px;padding:16px 24px;text-align:center;flex:1">
                    <div style="font-size:12px;color:#888">是否用券</div>
                    <div style="font-size:22px;font-weight:700">{'✅ 是' if traj.get('use_coupon') else '❌ 否'}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            # CoT 步骤式可视化
            st.markdown("**🧠 决策思维链 (Chain-of-Thought)**")
            thinking = traj.get("thinking", "")

            # 解析 thinking 中的 Step
            import re
            steps_raw = re.split(r'(Step\s*\d)', thinking)
            steps = []
            current = ""
            for part in steps_raw:
                if re.match(r'Step\s*\d', part):
                    if current.strip():
                        steps.append(current.strip())
                    current = part
                else:
                    current += part
            if current.strip():
                steps.append(current.strip())

            if len(steps) < 2:
                # 如果解析不出步骤，尝试用其他分隔符
                steps = [s.strip() for s in thinking.replace("Step4", "|Step4").replace("Step3", "|Step3").replace("Step2", "|Step2").replace("Step1", "|Step1").split("|") if s.strip()]

            step_colors = ["#1565c0", "#2e7d32", "#e65100", "#c62828"]
            step_icons = ["🔍", "🏪", "💰", "✅"]
            step_names = ["评估/需求判断", "选择/比较", "价格确认", "最终决策"]

            if steps:
                for i, step_text in enumerate(steps[:4]):
                    color = step_colors[i % len(step_colors)]
                    icon = step_icons[i % len(step_icons)]
                    name = step_names[i % len(step_names)]
                    st.markdown(f"""
                    <div style="display:flex;align-items:flex-start;margin-bottom:4px">
                        <div style="min-width:36px;height:36px;background:{color};border-radius:50%;display:flex;align-items:center;justify-content:center;color:white;font-size:16px;margin-right:12px;margin-top:2px">{icon}</div>
                        <div style="flex:1;background:#f8f9fa;border-left:3px solid {color};border-radius:0 8px 8px 0;padding:10px 14px">
                            <div style="font-size:11px;color:{color};font-weight:600;margin-bottom:2px">{name}</div>
                            <div style="font-size:13px;color:#333;line-height:1.5">{step_text}</div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    if i < len(steps) - 1:
                        st.markdown(f'<div style="margin-left:16px;color:#ccc;font-size:16px">↓</div>', unsafe_allow_html=True)
            else:
                st.info(thinking if thinking else "无思维链数据")

            st.markdown(f"""
            <div style="background:#e8f5e9;border-radius:8px;padding:10px 14px;margin-top:8px">
                <span style="font-size:12px;color:#2e7d32">💡 <b>决策原因:</b></span>
                <span style="font-size:13px"> {traj.get('reason', '无')}</span>
            </div>
            """, unsafe_allow_html=True)


            # ── 全场景对比 ──
            st.markdown("#### 📊 该用户全场景行为对比")
            table_html = """
            <table class="strategy-table" style="table-layout:fixed;width:100%">
                <colgroup><col style="width:35%"><col style="width:20%"><col style="width:20%"><col style="width:25%"></colgroup>
                <tr><th style="text-align:left">场景</th><th>行为</th><th>置信度</th><th>用券</th></tr>
            """
            for sk, tr in scenes.items():
                sc_info = SCENARIOS.get(sk, {})
                a_html = {"order":'<span class="tag-green">下单</span>',"cart":'<span class="tag-yellow">加购</span>',
                          "browse":'<span class="tag-yellow">浏览</span>',"ignore":'<span class="tag-red">忽略</span>'}
                coupon_html = '<span class="tag-green">是</span>' if tr.get("use_coupon") else '<span style="color:#999">否</span>'
                table_html += f'<tr><td style="text-align:left">{sc_info.get("icon","")} {sc_info.get("label",sk)}</td><td>{a_html.get(tr["action"],tr["action"])}</td><td>{tr["confidence"]:.2f}</td><td>{coupon_html}</td></tr>'
            table_html += "</table>"
            st.markdown(table_html, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# 页面 4: 风险与建议
# ═══════════════════════════════════════════════════════════════
with tab_risk:
    st.markdown("## ⚠️ 风险预警与策略建议")

    tab1, tab2, tab3 = st.tabs(["🚨 风险预警", "💡 业务洞察", "📋 投放策略"])

    with tab1:
        st.markdown("""
        <div class="alert-red">
            <strong>🚨 品质导向用户：全场景发券亏损</strong><br>
            <span style="font-size:13px">26 名品质导向用户在所有 7 个有券场景中 net_value 均为负（-10 ~ -27 元/单）。<br>
            <b>根因：</b>客单价高→核销概率高→成本高，但购买决策由品质驱动，券的增量转化极小。<br>
            <b>建议：</b>停止对品质用户发满减券，改用品质推荐、新店体验等非价格策略。</span>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="alert-red">
            <strong>🚨 暴雨场景：盲目发券加剧供需失衡</strong><br>
            <span style="font-size:13px">周末暴雨晚餐场景正向占比仅 56%，是所有场景最低。满 50 门槛过高，低客单价用户够不到。<br>
            <b>建议：</b>暴雨场景改为"免配送费"，降低决策门槛且不过度刺激需求。</span>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="alert-orange">
            <strong>⚠️ 刚需用户：部分场景利润倒贴</strong><br>
            <span style="font-size:13px">刚需用户在暴雨晚餐（-2.60元）、聚餐（-2.51元）、夜宵（-2.04元）场景 net_value 为负。<br>
            <b>根因：</b>这些场景下刚需用户本来就会下单，发券是纯利润倒贴。<br>
            <b>建议：</b>对刚需用户在这三个场景停止发券，预算转投薅羊毛党。</span>
        </div>
        """, unsafe_allow_html=True)

    with tab2:
        st.markdown("""
        <div class="alert-green">
            <strong>💡 工作日午餐：最安全的发券场景</strong><br>
            <span style="font-size:13px">正向占比 <b>81.5%</b>，平均 uplift +0.461，平均 net_value +20.31 元。<br>
            用户基数大、需求稳定，券的边际效果最可靠。</span>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="alert-green">
            <strong>💡 节假日聚餐：uplift_gmv 最高</strong><br>
            <span style="font-size:13px">平均 uplift_gmv <b>+31.08 元</b>，正向占比 76%。<br>
            高客单价放大增量，一张 ¥25 券可撬动 ¥30+ 增量 GMV（对非品质用户）。</span>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="alert-green">
            <strong>💡 薅羊毛党：精准投放 ROI 最高</strong><br>
            <span style="font-size:13px">虽然整体正向占比仅 41.4%，但在高折扣场景（暴雨专享 40%、下午茶 32%）uplift 极大。<br>
            <b>策略：</b>精准识别 + 匹配高折扣率场景，集中火力投放。</span>
        </div>
        """, unsafe_allow_html=True)

    with tab3:
        st.markdown("#### 分人群 × 分场景投放矩阵")
        st.markdown("""
        <table class="strategy-table" style="table-layout:fixed;width:100%">
            <colgroup>
                <col style="width:14%">
                <col style="width:11%"><col style="width:11%"><col style="width:11%"><col style="width:11%">
                <col style="width:14%"><col style="width:14%"><col style="width:14%">
            </colgroup>
            <tr>
                <th>用户类型</th>
                <th>🍱 午餐</th>
                <th>🧋 下午茶</th>
                <th>🥐 早餐</th>
                <th>⛈️ 暴雨</th>
                <th>🌙 夜宵</th>
                <th>🎉 聚餐</th>
                <th>🌧️ 暴雨晚餐</th>
            </tr>
            <tr>
                <td><b>🟢 刚需用户</b></td>
                <td><span class="tag-green">✅ 发券</span></td>
                <td><span class="tag-green">✅ 发券</span></td>
                <td><span class="tag-green">✅ 发券</span></td>
                <td><span class="tag-green">✅ 发券</span></td>
                <td><span class="tag-yellow">⚠️ 谨慎</span></td>
                <td><span class="tag-yellow">⚠️ 谨慎</span></td>
                <td><span class="tag-yellow">⚠️ 谨慎</span></td>
            </tr>
            <tr>
                <td><b>🔴 品质导向</b></td>
                <td><span class="tag-red">❌ 不发</span></td>
                <td><span class="tag-red">❌ 不发</span></td>
                <td><span class="tag-red">❌ 不发</span></td>
                <td><span class="tag-red">❌ 不发</span></td>
                <td><span class="tag-red">❌ 不发</span></td>
                <td><span class="tag-red">❌ 不发</span></td>
                <td><span class="tag-red">❌ 不发</span></td>
            </tr>
            <tr>
                <td><b>🟡 薅羊毛党</b></td>
                <td><span class="tag-green">✅ 发券</span></td>
                <td><span class="tag-green">✅ 发券</span></td>
                <td><span class="tag-yellow">⚠️ 看折扣</span></td>
                <td><span class="tag-green">✅ 发券</span></td>
                <td><span class="tag-yellow">⚠️ 看习惯</span></td>
                <td><span class="tag-red">❌ 门槛高</span></td>
                <td><span class="tag-yellow">⚠️ 看客单</span></td>
            </tr>
        </table>
        """, unsafe_allow_html=True)

        st.markdown("")
        st.markdown("#### 预算分配建议（budget=¥300）")

        b1, b2, b3, b4 = st.columns(4)
        b1.markdown("""
        <div class="card" style="text-align:center">
            <div style="font-size:28px;font-weight:700;color:#2e7d32">60%</div>
            <div style="font-size:12px;color:#888">¥180</div>
            <div style="font-size:13px;margin-top:8px">刚需 × 午餐/下午茶/早餐/暴雨</div>
            <div style="font-size:11px;color:#999">最稳定的正向增量</div>
        </div>
        """, unsafe_allow_html=True)
        b2.markdown("""
        <div class="card" style="text-align:center">
            <div style="font-size:28px;font-weight:700;color:#f57c00">30%</div>
            <div style="font-size:12px;color:#888">¥90</div>
            <div style="font-size:13px;margin-top:8px">薅羊毛党 × 高折扣场景</div>
            <div style="font-size:11px;color:#999">uplift 最大，需精准匹配</div>
        </div>
        """, unsafe_allow_html=True)
        b3.markdown("""
        <div class="card" style="text-align:center">
            <div style="font-size:28px;font-weight:700;color:#1565c0">10%</div>
            <div style="font-size:12px;color:#888">¥30</div>
            <div style="font-size:13px;margin-top:8px">预留突发场景</div>
            <div style="font-size:11px;color:#999">极端天气临时加码</div>
        </div>
        """, unsafe_allow_html=True)
        b4.markdown("""
        <div class="card" style="text-align:center">
            <div style="font-size:28px;font-weight:700;color:#c62828">0%</div>
            <div style="font-size:12px;color:#888">¥0</div>
            <div style="font-size:13px;margin-top:8px">品质导向用户</div>
            <div style="font-size:11px;color:#999">全场景亏损，改用非价格策略</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("")
        st.markdown("#### 替代策略")
        a1, a2 = st.columns(2)
        with a1:
            st.markdown("""
            <div class="card">
                <b>🔴 品质导向用户替代方案</b>
                <ul style="font-size:13px;margin-top:8px">
                    <li>🏪 新店体验券（品质背书）</li>
                    <li>⭐ 高评分店铺专属推荐</li>
                    <li>🎁 会员专属权益（非价格驱动）</li>
                </ul>
            </div>
            """, unsafe_allow_html=True)
        with a2:
            st.markdown("""
            <div class="card">
                <b>🌧️ 暴雨场景替代方案</b>
                <ul style="font-size:13px;margin-top:8px">
                    <li>🚚 免配送费（降低决策门槛）</li>
                    <li>⏰ 预售膨胀券（锁定沉没成本）</li>
                    <li>📱 Push 提醒而非主动发券</li>
                </ul>
            </div>
            """, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# 页面 5: 实时仿真
# ═══════════════════════════════════════════════════════════════

# 概率引擎（复用 cold_boot_simulation.py 的逻辑）
SIM_SCENARIOS = [
    {"key": "weekday_lunch", "time": "工作日 12:00", "weather": "晴", "coupon_threshold": 30, "coupon_discount": 8, "is_meal_time": True, "is_weekend": False, "is_bad_weather": False},
    {"key": "afternoon_tea", "time": "工作日 15:00", "weather": "小雨", "coupon_threshold": 25, "coupon_discount": 8, "is_meal_time": False, "is_weekend": False, "is_bad_weather": True},
    {"key": "weekend_rain_dinner", "time": "周末 18:00", "weather": "暴雨", "coupon_threshold": 50, "coupon_discount": 15, "is_meal_time": True, "is_weekend": True, "is_bad_weather": True},
    {"key": "weekday_breakfast", "time": "工作日 08:00", "weather": "晴", "coupon_threshold": 15, "coupon_discount": 5, "is_meal_time": True, "is_weekend": False, "is_bad_weather": False},
    {"key": "weekend_no_coupon", "time": "周末 11:00", "weather": "晴", "coupon_threshold": 0, "coupon_discount": 0, "is_meal_time": True, "is_weekend": True, "is_bad_weather": False},
    {"key": "late_night", "time": "工作日 22:00", "weather": "晴", "coupon_threshold": 20, "coupon_discount": 6, "is_meal_time": False, "is_weekend": False, "is_bad_weather": False},
    {"key": "holiday_gathering", "time": "节假日 12:00", "weather": "晴", "coupon_threshold": 80, "coupon_discount": 25, "is_meal_time": True, "is_weekend": True, "is_bad_weather": False},
    {"key": "rain_stuck", "time": "工作日 16:00", "weather": "暴雨", "coupon_threshold": 30, "coupon_discount": 12, "is_meal_time": False, "is_weekend": False, "is_bad_weather": True},
]

import random
random.seed(42)

def sim_compute_probs(user, scenario):
    basket = user["avg_basket_size"]
    elasticity = user["elasticity_type"]
    impulse = user["impulse_buy_rate"]
    compare = user["deep_compare_rate"]
    health = user.get("lifecycle_health_score", 0.5)
    late_night = user.get("late_night_actions", 0)
    bad_weather_ratio = user.get("bad_weather_order_ratio", 0.15)
    discount_threshold = user.get("avg_discount_enjoyed", 0.25)

    base = 0.25 + health * 0.15
    cd = scenario["coupon_discount"]
    ct = scenario["coupon_threshold"]

    if cd > 0:
        dr = cd / max(ct, 1)
        if elasticity == "高弹性(薅羊毛党)":
            base += 0.20 if dr >= discount_threshold else -0.10
        elif elasticity == "低弹性(撇脂定价目标)":
            base += 0.05
        else:
            base += 0.08
    else:
        if elasticity == "高弹性(薅羊毛党)":
            base -= 0.15
        elif elasticity == "缺乏弹性(刚需)":
            base += 0.05

    ep = (ct - cd) / max(basket, 1) if cd > 0 else 1.0
    if ep < 0.7: base += 0.10
    elif ep > 1.5: base -= 0.15
    elif ep > 1.3: base -= 0.08

    if scenario["is_meal_time"]: base += 0.10
    else: base -= 0.05

    if "22:00" in scenario["time"]:
        if late_night > 2: base += 0.10
        elif late_night == 0: base -= 0.20

    if scenario["is_bad_weather"]:
        if bad_weather_ratio > 0.3: base += 0.12
        elif bad_weather_ratio > 0.1: base += 0.05
        else: base -= 0.05

    base += impulse * 0.3
    base -= compare * 0.2

    order_p = max(0.03, min(0.85, base))
    rem = 1.0 - order_p
    cart_p = rem * 0.30
    browse_p = rem * 0.35
    ignore_p = rem - cart_p - browse_p
    return order_p, cart_p, browse_p, max(0, ignore_p)

def sim_gen_confidence(user, action):
    imp = user["impulse_buy_rate"]
    cmp = user["deep_compare_rate"]
    if action == "order": return round(max(0.3, min(0.95, 0.55 + imp*1.5 - cmp*0.8 + random.gauss(0,0.08))), 2)
    elif action == "cart": return round(max(0.15, min(0.55, 0.30 + cmp*0.5 + random.gauss(0,0.06))), 2)
    elif action == "browse": return round(max(0.10, min(0.40, 0.25 + random.gauss(0,0.06))), 2)
    else: return round(max(0.02, min(0.15, 0.08 + random.gauss(0,0.03))), 2)

def sim_gen_thinking(user, action, scenario):
    e = user["elasticity_type"]
    basket = user["avg_basket_size"]
    cd = scenario["coupon_discount"]
    ct = scenario["coupon_threshold"]

    if e == "高弹性(薅羊毛党)":
        if cd > 0:
            rate = cd / max(ct,1) * 100
            thr = user.get("avg_discount_enjoyed", 0.25) * 100
            eff = ct - cd
            t = f"Step1-券值评估: 减{cd}元, 折扣率{rate:.0f}%, 我的阈值{thr:.0f}%. "
            if rate >= thr:
                t += f"达标。Step2-凑单: 凑到{ct}元, 实付{eff}元, "
                t += "划算。" if eff < basket*0.7 else ("舒适区。" if eff < basket*1.3 else "偏贵。")
                t += "Step3-替代比较: 没找到更好的。"
            else:
                t += "力度不够。Step2: 跳过。Step3: 再等等。"
        else:
            t = "Step1: 没有券, 我一般有券才下单。再等等。"
    elif e == "低弹性(撇脂定价目标)":
        t = "Step1-需求: " + ("到饭点了, 想吃好的。" if scenario["is_meal_time"] else "不是饭点, 看看。")
        t += "Step2-品牌: 看重品质, 评分低的不考虑。"
        t += f"Step3-价格: {'有券减'+str(cd)+'元, 顺手用。' if cd > 0 else '没券也无所谓。'}"
    else:
        t = "Step1-需求: " + ("饿了, 到饭点了。" if scenario["is_meal_time"] else "不是饭点, 没特别想吃的。")
        t += "Step2-习惯: 看看常去的店。"
        t += f"Step3-价格: {'有券就用, 减'+str(cd)+'块。' if cd > 0 else '没券也行, 该吃就吃。'}"

    action_desc = {"order": "Step4: 决定下单。", "cart": "Step4: 先加购, 等会再说。", "browse": "Step4: 再看看。", "ignore": "Step4: 算了, 不需要。"}
    t += action_desc.get(action, "")
    return t


with tab_realtime:
    st.markdown("## 🧪 实时仿真 — 输入用户画像，预测消费行为")

    col_input, col_result = st.columns([1, 2])

    with col_input:
        st.markdown("#### 用户画像输入")

        # 快速模板
        template = st.selectbox("快速模板", ["自定义", "典型刚需用户", "典型薅羊毛党", "典型品质用户"])
        if template == "典型刚需用户":
            defaults = {"basket": 25.0, "etype": "缺乏弹性(刚需)", "impulse": 0.10, "compare": 0.08, "night": 1, "weather": 0.15, "hhi": 0.45, "health": 0.6, "discount": 0.20}
        elif template == "典型薅羊毛党":
            defaults = {"basket": 35.0, "etype": "高弹性(薅羊毛党)", "impulse": 0.05, "compare": 0.12, "night": 0, "weather": 0.10, "hhi": 0.25, "health": 0.4, "discount": 0.26}
        elif template == "典型品质用户":
            defaults = {"basket": 120.0, "etype": "低弹性(撇脂定价目标)", "impulse": 0.08, "compare": 0.06, "night": 0, "weather": 0.20, "hhi": 0.55, "health": 0.7, "discount": 0.15}
        else:
            defaults = {"basket": 40.0, "etype": "缺乏弹性(刚需)", "impulse": 0.10, "compare": 0.10, "night": 1, "weather": 0.15, "hhi": 0.35, "health": 0.5, "discount": 0.20}

        basket = st.slider("客单价 (元)", 10.0, 500.0, defaults["basket"], 5.0)
        etype = st.selectbox("弹性类型", ["缺乏弹性(刚需)", "低弹性(撇脂定价目标)", "高弹性(薅羊毛党)"],
                             index=["缺乏弹性(刚需)", "低弹性(撇脂定价目标)", "高弹性(薅羊毛党)"].index(defaults["etype"]))
        impulse = st.slider("冲动下单率", 0.0, 0.30, defaults["impulse"], 0.01)
        compare = st.slider("深度比价率", 0.0, 0.30, defaults["compare"], 0.01)
        night = st.slider("夜宵行为次数", 0, 10, defaults["night"])
        weather = st.slider("恶劣天气下单占比", 0.0, 0.60, defaults["weather"], 0.05)
        hhi = st.slider("品牌集中度 (HHI)", 0.0, 1.0, defaults["hhi"], 0.05)
        health = st.slider("活跃健康度", 0.0, 1.0, defaults["health"], 0.05)
        discount_thr = st.slider("心理折扣阈值", 0.0, 0.50, defaults["discount"], 0.01)

        run_sim = st.button("🚀 开始仿真", use_container_width=True, type="primary")

    with col_result:
        if run_sim:
            sim_user = {
                "avg_basket_size": basket,
                "elasticity_type": etype,
                "impulse_buy_rate": impulse,
                "deep_compare_rate": compare,
                "late_night_actions": night,
                "bad_weather_order_ratio": weather,
                "brand_hhi": hhi,
                "lifecycle_health_score": health,
                "avg_discount_enjoyed": discount_thr,
                "avg_add_cart_per_session": 0.3,
                "high_discount_hunter_ratio": 0.5 if etype == "高弹性(薅羊毛党)" else 0.15,
                "weekend_poi_explore": 10,
            }

            # 无券基线
            ctrl_sc = [s for s in SIM_SCENARIOS if s["key"] == "weekend_no_coupon"][0]
            ctrl_probs = sim_compute_probs(sim_user, ctrl_sc)
            ctrl_action = ["order","cart","browse","ignore"][ctrl_probs.index(max(ctrl_probs))]
            ctrl_p = ctrl_probs[0]  # order prob

            st.markdown("#### 仿真结果")
            st.caption(f"用户画像: {etype} | 客单价 ¥{basket:.0f} | 冲动率 {impulse:.0%} | 比价率 {compare:.0%}")
            st.markdown("")

            # 每个场景的结果 — 一次性拼接完整表格
            best_scene = None
            best_uplift = -999

            sim_table = """
            <table class="strategy-table" style="table-layout:fixed;width:100%">
                <colgroup>
                    <col style="width:22%"><col style="width:14%"><col style="width:14%">
                    <col style="width:14%"><col style="width:14%"><col style="width:14%">
                </colgroup>
                <tr>
                    <th style="text-align:left">场景</th>
                    <th>预测行为</th>
                    <th>下单概率</th>
                    <th>置信度</th>
                    <th>Uplift</th>
                    <th>建议</th>
                </tr>
            """

            for sc in SIM_SCENARIOS:
                if sc["key"] == "weekend_no_coupon":
                    continue
                probs = sim_compute_probs(sim_user, sc)
                actions = ["order", "cart", "browse", "ignore"]
                action = actions[probs.index(max(probs))]
                conf = sim_gen_confidence(sim_user, action)
                order_p = probs[0]
                uplift = order_p - ctrl_p

                if uplift > best_uplift:
                    best_uplift = uplift
                    best_scene = sc["key"]

                action_html = {"order": '<span class="tag-green">✅ 下单</span>',
                               "cart": '<span class="tag-yellow">🛒 加购</span>',
                               "browse": '<span class="tag-yellow">👀 浏览</span>',
                               "ignore": '<span class="tag-red">💤 忽略</span>'}
                uplift_color = "#2e7d32" if uplift > 0.05 else ("#e65100" if uplift > 0 else "#c62828")
                advice = '<span class="tag-green">✅ 发券</span>' if uplift > 0.05 else ('<span class="tag-yellow">⚠️ 谨慎</span>' if uplift > 0 else '<span class="tag-red">❌ 不发</span>')

                sc_info = SCENARIOS.get(sc["key"], {})
                order_p_str = f"{order_p:.1%}"
                uplift_str = f"{uplift:+.1%}"
                sim_table += (
                    "<tr>"
                    f'<td style="text-align:left">{sc_info.get("icon","")} {sc_info.get("label","")}<br>'
                    f'<span style="font-size:10px;color:#999">{sc_info.get("coupon","")}</span></td>'
                    f"<td>{action_html.get(action, action)}</td>"
                    f"<td>{order_p_str}</td>"
                    f"<td>{conf}</td>"
                    f'<td style="color:{uplift_color};font-weight:600">{uplift_str}</td>'
                    f"<td>{advice}</td>"
                    "</tr>"
                )

            sim_table += "</table>"
            st.markdown(sim_table, unsafe_allow_html=True)

            # 无券基线
            st.caption(f"📊 无券基线下单概率: {ctrl_p:.1%} | 最佳发券场景: {SCENARIOS.get(best_scene,{}).get('label','')} (uplift {best_uplift:+.1%})")

            # 用户分群
            if best_uplift > 0.05 and ctrl_p < 0.3:
                segment = "可说服用户 — 无券不买，有券才买，发券价值最高"
                seg_color = "#2e7d32"
            elif abs(best_uplift) <= 0.05 and ctrl_p >= 0.3:
                segment = "自然转化用户 — 发不发券都会买，不建议发券"
                seg_color = "#1565c0"
            elif abs(best_uplift) <= 0.05 and ctrl_p < 0.3:
                segment = "沉睡用户 — 发不发券都不买，发券无效"
                seg_color = "#999"
            else:
                segment = "锦上添花用户 — 有一定自然转化，券能进一步提升"
                seg_color = "#f57c00"

            st.markdown(f'<div style="background:#f5f5f5;border-left:4px solid {seg_color};padding:12px 16px;border-radius:6px;margin-top:12px"><b>用户分群:</b> {segment}</div>', unsafe_allow_html=True)

            # CoT 展示
            if best_scene:
                st.markdown(f"#### 🧠 最佳场景决策思维链")
                best_sc_data = [s for s in SIM_SCENARIOS if s["key"] == best_scene][0]
                best_probs = sim_compute_probs(sim_user, best_sc_data)
                best_action = ["order","cart","browse","ignore"][best_probs.index(max(best_probs))]
                thinking = sim_gen_thinking(sim_user, best_action, best_sc_data)
                st.info(thinking)
        else:
            st.markdown("""
            <div style="background:linear-gradient(135deg,#f8f9fa,#e8eaf6);border-radius:12px;padding:40px 30px;text-align:center">
                <div style="font-size:48px;margin-bottom:16px">🧪</div>
                <div style="font-size:18px;font-weight:600;color:#333;margin-bottom:8px">在左侧输入用户画像</div>
                <div style="font-size:13px;color:#888;margin-bottom:20px">选择快速模板或自定义参数，点击「开始仿真」查看预测结果</div>
                <div style="display:flex;justify-content:center;gap:24px;flex-wrap:wrap">
                    <div style="background:white;border-radius:8px;padding:14px 18px;min-width:140px;box-shadow:0 1px 4px rgba(0,0,0,0.06)">
                        <div style="font-size:20px">🎯</div>
                        <div style="font-size:12px;font-weight:600;color:#1565c0;margin-top:4px">7 个场景预测</div>
                        <div style="font-size:11px;color:#999">午餐/下午茶/暴雨/夜宵...</div>
                    </div>
                    <div style="background:white;border-radius:8px;padding:14px 18px;min-width:140px;box-shadow:0 1px 4px rgba(0,0,0,0.06)">
                        <div style="font-size:20px">📊</div>
                        <div style="font-size:12px;font-weight:600;color:#2e7d32;margin-top:4px">Uplift 分析</div>
                        <div style="font-size:11px;color:#999">发券 vs 不发券的增量</div>
                    </div>
                    <div style="background:white;border-radius:8px;padding:14px 18px;min-width:140px;box-shadow:0 1px 4px rgba(0,0,0,0.06)">
                        <div style="font-size:20px">🧠</div>
                        <div style="font-size:12px;font-weight:600;color:#e65100;margin-top:4px">决策思维链</div>
                        <div style="font-size:11px;color:#999">CoT 可解释推理过程</div>
                    </div>
                    <div style="background:white;border-radius:8px;padding:14px 18px;min-width:140px;box-shadow:0 1px 4px rgba(0,0,0,0.06)">
                        <div style="font-size:20px">👥</div>
                        <div style="font-size:12px;font-weight:600;color:#c62828;margin-top:4px">用户分群</div>
                        <div style="font-size:11px;color:#999">可说服/自然转化/沉睡</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
